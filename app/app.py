"""
Pashto Whisper STT — Deployable Streamlit App
==============================================
Runs locally on a user's device (VS Code / terminal: `streamlit run app.py`).
No Google Drive dependency — all data lives under a local `workspace_data/`
folder next to this file (override with the APP_DATA_DIR env var).

Transcription is delegated to the project's FastAPI service (api/) over
HTTP — this app no longer loads the Whisper model in-process. Start the API
separately (see README) before using the Transcribe tab. Translation
(NLLB) and everything else still run locally in this Streamlit process,
unchanged.
"""

import csv
import hashlib
import io
import os
import re
import secrets
import sqlite3
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import streamlit as st
import torch

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import yaml
except ImportError:
    yaml = None

try:
    from jiwer import wer as jiwer_wer
    JIWER_AVAILABLE = True
except ImportError:
    JIWER_AVAILABLE = False

import requests
from transformers import pipeline as hf_pipeline

APP_DIR = Path(__file__).resolve().parent
ROOT_DIR = APP_DIR.parent

# ── Config file (configs/config.yaml) ──────────────────────────────────────
CONFIG_PATH = ROOT_DIR / "configs" / "config.yaml"


def load_config():
    if yaml is None or not CONFIG_PATH.exists():
        return {}
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


CONFIG = load_config()

# ── Transcription API client config ────────────────────────────────────────
# Matches api/config.py's Settings.API_V1_STR default ("/api/v1") and the
# router prefix in api/routes/transcription.py ("/transcription").
API_BASE_URL = os.environ.get("API_BASE_URL", "http://localhost:8000").rstrip("/")
API_V1_STR = os.environ.get("API_V1_STR", "/api/v1")
API_TRANSCRIPTION_PREFIX = f"{API_V1_STR}/transcription"

# FIX: Point health directly to the root application space as defined in main.py
API_HEALTH_URL = f"{API_BASE_URL}/health"

# These stay the same as they correctly process through the transcription router
API_MODEL_INFO_URL = f"{API_BASE_URL}{API_TRANSCRIPTION_PREFIX}/model-info"
API_TRANSCRIBE_URL = f"{API_BASE_URL}{API_TRANSCRIPTION_PREFIX}/transcribe"
API_REQUEST_TIMEOUT = int(os.environ.get("API_REQUEST_TIMEOUT", "600"))


# ── Local, per-device storage (no Drive dependency) ────────────────────────
APP_DATA_DIR = Path(os.environ.get("APP_DATA_DIR", APP_DIR / "workspace_data")).resolve()
AUDIO_STORE = APP_DATA_DIR / "audio_uploads"
DB_PATH = APP_DATA_DIR / "pashto_app.db"
APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
AUDIO_STORE.mkdir(parents=True, exist_ok=True)

SUPPORTED_FORMATS = ["wav", "mp3", "mp4", "m4a", "flac", "ogg", "opus", "webm", "aac", "wma"]

NLLB_MODEL = "facebook/nllb-200-distilled-600M"
PASHTO_CODE = "pbt_Arab"
ENGLISH_CODE = "eng_Latn"
URDU_CODE = "urd_Arab"

# AI Assistant hits external free inference APIs with your transcription text
# as context. Always visible in the UI; each user supplies their own free HF
# token the first time they open the tab, and it's saved to their account.

# ── Database ────────────────────────────────────────────────────────────────

def get_db():
    return sqlite3.connect(str(DB_PATH), check_same_thread=False)


def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "CREATE TABLE IF NOT EXISTS users ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "username TEXT UNIQUE NOT NULL,"
        "email TEXT UNIQUE NOT NULL,"
        "password_salt TEXT NOT NULL,"
        "password_hash TEXT NOT NULL,"
        "created_at TEXT)"
    )
    c.execute(
        "CREATE TABLE IF NOT EXISTS transcriptions ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "user_id INTEGER NOT NULL,"
        "filename TEXT NOT NULL,"
        "original_transcription TEXT,"
        "edited_transcription TEXT,"
        "reference_text TEXT,"
        "wer_score REAL,"
        "translation_en TEXT,"
        "translation_ur TEXT,"
        "audio_filepath TEXT,"
        "created_at TEXT,"
        "FOREIGN KEY(user_id) REFERENCES users(id))"
    )
    conn.commit()

    # Lightweight migrations for older DBs created by the previous app.py
    user_cols = {row[1] for row in c.execute("PRAGMA table_info(users)")}
    if "email" not in user_cols:
        c.execute("ALTER TABLE users ADD COLUMN email TEXT")
    if "password_salt" not in user_cols:
        c.execute("ALTER TABLE users ADD COLUMN password_salt TEXT")
    if "ai_hf_token" not in user_cols:
        c.execute("ALTER TABLE users ADD COLUMN ai_hf_token TEXT")
    if "ai_hf_token_updated_at" not in user_cols:
        c.execute("ALTER TABLE users ADD COLUMN ai_hf_token_updated_at TEXT")

    trans_cols = {row[1] for row in c.execute("PRAGMA table_info(transcriptions)")}
    migrations = {
        "translation_en": "ALTER TABLE transcriptions ADD COLUMN translation_en TEXT",
        "translation_ur": "ALTER TABLE transcriptions ADD COLUMN translation_ur TEXT",
        "wer_score": "ALTER TABLE transcriptions ADD COLUMN wer_score REAL",
        "reference_text": "ALTER TABLE transcriptions ADD COLUMN reference_text TEXT",
        "edited_transcription": "ALTER TABLE transcriptions ADD COLUMN edited_transcription TEXT",
        "audio_filepath": "ALTER TABLE transcriptions ADD COLUMN audio_filepath TEXT",
    }
    for col, sql in migrations.items():
        if col not in trans_cols:
            c.execute(sql)
    conn.commit()
    conn.close()


init_db()

# ── Auth: salted PBKDF2 hashing, username+email+password validation ───────

USERNAME_RE = re.compile(r"^[A-Za-z0-9_]{3,30}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _hash_password(password, salt_hex=None):
    if salt_hex is None:
        salt_hex = secrets.token_hex(16)
    pw_hash = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), bytes.fromhex(salt_hex), 100_000
    ).hex()
    return salt_hex, pw_hash


def _verify_password(password, salt_hex, expected_hash):
    if not salt_hex:
        # Legacy accounts created by the old unsalted-sha256 scheme
        return secrets.compare_digest(
            hashlib.sha256(password.encode()).hexdigest(), expected_hash
        )
    _, pw_hash = _hash_password(password, salt_hex)
    return secrets.compare_digest(pw_hash, expected_hash)


def validate_signup(username, email, password, confirm):
    if not USERNAME_RE.match(username):
        return "Username must be 3-30 characters: letters, numbers, underscore only."
    if not EMAIL_RE.match(email):
        return "Please enter a valid email address."
    if len(password) < 8:
        return "Password must be at least 8 characters."
    if not re.search(r"[A-Za-z]", password) or not re.search(r"[0-9]", password):
        return "Password must contain both letters and numbers."
    if password != confirm:
        return "Passwords do not match."
    return None


def create_user(username, email, password):
    conn = get_db()
    c = conn.cursor()
    if c.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        conn.close()
        return False, "Username already taken."
    if c.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone():
        conn.close()
        return False, "An account with this email already exists."
    salt, pw_hash = _hash_password(password)
    c.execute(
        "INSERT INTO users (username, email, password_salt, password_hash, created_at) "
        "VALUES (?, ?, ?, ?, ?)",
        (username, email, salt, pw_hash, datetime.now().isoformat()),
    )
    conn.commit()
    conn.close()
    return True, "Account created! Please log in."


def get_user_ai_token(user_id):
    conn = get_db()
    row = conn.execute("SELECT ai_hf_token FROM users WHERE id=?", (user_id,)).fetchone()
    conn.close()
    return row[0].strip() if row and row[0] else None


def save_user_ai_token(user_id, token):
    conn = get_db()
    conn.execute(
        "UPDATE users SET ai_hf_token=?, ai_hf_token_updated_at=? WHERE id=?",
        (token.strip(), datetime.now().isoformat(), user_id),
    )
    conn.commit()
    conn.close()


def clear_user_ai_token(user_id):
    conn = get_db()
    conn.execute("UPDATE users SET ai_hf_token=NULL, ai_hf_token_updated_at=NULL WHERE id=?", (user_id,))
    conn.commit()
    conn.close()


def verify_user(identifier, password):
    """identifier can be a username or an email address."""
    conn = get_db()
    row = conn.execute(
        "SELECT id, password_salt, password_hash, username FROM users "
        "WHERE username=? OR email=?",
        (identifier, identifier),
    ).fetchone()
    conn.close()
    if row and _verify_password(password, row[1], row[2]):
        return True, row[0], row[3]
    return False, None, None


# ── Transcription data access ───────────────────────────────────────────────

def save_transcription(user_id, filename, original, audio_filepath=None, reference=None, wer_score=None):
    conn = get_db()
    c = conn.cursor()
    existing = c.execute(
        "SELECT id FROM transcriptions "
        "WHERE user_id=? AND filename=? AND original_transcription=? AND (audio_filepath=? OR audio_filepath IS NULL)",
        (user_id, filename, original, audio_filepath),
    ).fetchone()
    if existing:
        conn.close()
        return existing[0]
    c.execute(
        "INSERT INTO transcriptions "
        "(user_id, filename, original_transcription, edited_transcription, "
        "reference_text, wer_score, translation_en, translation_ur, audio_filepath, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, filename, original, original, reference, wer_score, None, None, audio_filepath, datetime.now().isoformat()),
    )
    conn.commit()
    tid = c.lastrowid
    conn.close()
    return tid


def update_transcription(tid, edited, reference=None, wer_score=None):
    conn = get_db()
    conn.execute(
        "UPDATE transcriptions SET edited_transcription=?, reference_text=?, wer_score=? WHERE id=?",
        (edited, reference, wer_score, tid),
    )
    conn.commit()
    conn.close()


def delete_transcription(tid, audio_filepath=None):
    conn = get_db()
    conn.execute("DELETE FROM transcriptions WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    if audio_filepath and os.path.exists(audio_filepath):
        try:
            os.remove(audio_filepath)
            parent_dir = os.path.dirname(audio_filepath)
            if parent_dir != str(AUDIO_STORE) and not os.listdir(parent_dir):
                os.rmdir(parent_dir)
        except OSError as e:
            st.warning(f"Could not delete audio file: {e}")


def save_translations(tid, en_text, ur_text):
    conn = get_db()
    conn.execute(
        "UPDATE transcriptions SET translation_en=?, translation_ur=? WHERE id=?",
        (en_text, ur_text, tid),
    )
    conn.commit()
    conn.close()


def get_user_history(user_id):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, filename, original_transcription, edited_transcription, "
        "reference_text, wer_score, translation_en, translation_ur, audio_filepath, created_at "
        "FROM transcriptions WHERE user_id=? ORDER BY created_at DESC",
        (user_id,),
    ).fetchall()
    conn.close()
    return rows

# ── Transcription — delegated to the FastAPI service over HTTP ────────────


@st.cache_data(ttl=30, show_spinner=False)
def check_api_health():
    """Cached for 30s so we don't hit /health on every Streamlit rerun."""
    try:
        r = requests.get(API_HEALTH_URL, timeout=50)
        if r.status_code == 200:
            return True, r.json().get("status", "healthy")
        return False, f"HTTP {r.status_code}"
    except requests.exceptions.RequestException as e:
        return False, str(e)


@st.cache_data(ttl=30, show_spinner=False)
def get_api_model_info():
    try:
        r = requests.get(API_MODEL_INFO_URL, timeout=50)
        if r.status_code == 200:
            return r.json()
    except requests.exceptions.RequestException:
        pass
    return None


def transcribe_via_api(audio_path, reference_text=None):
    """
    Calls POST {API_TRANSCRIBE_URL} with the audio file (multipart) and an
    optional reference_text form field, matching api/routes/transcription.py
    and api/schemas/response.py::TranscriptionResponse.

    Returns (transcription, wer_score, processing_time_sec, model_version, error).
    On failure, the first four values are None and `error` holds a
    user-facing message.
    """
    filename = os.path.basename(audio_path)
    try:
        with open(audio_path, "rb") as f:
            files = {"file": (filename, f)}
            data = {}
            if reference_text and reference_text.strip():
                data["reference_text"] = reference_text.strip()
            resp = requests.post(API_TRANSCRIBE_URL, files=files, data=data, timeout=API_REQUEST_TIMEOUT)
    except requests.exceptions.ConnectionError:
        return None, None, None, None, (
            f"Could not reach the Transcription API at {API_BASE_URL}. "
            "Make sure it's running, e.g.:\n\n`uvicorn api.main:app --host 0.0.0.0 --port 8000`"
        )
    except requests.exceptions.Timeout:
        return None, None, None, None, "The Transcription API timed out. Try a shorter audio clip or check the API server."
    except requests.exceptions.RequestException as e:
        return None, None, None, None, f"Request to the Transcription API failed: {e}"

    if resp.status_code == 200:
        payload = resp.json()
        return (
            payload.get("transcription", ""),
            payload.get("wer_score"),
            payload.get("processing_time_sec"),
            payload.get("model_version"),
            None,
        )

    # api/exceptions.py returns {"success": False, "detail": "..."} on errors
    try:
        detail = resp.json().get("detail", resp.text)
    except ValueError:
        detail = resp.text
    return None, None, None, None, f"API error ({resp.status_code}): {detail}"


@st.cache_resource(show_spinner="Loading translation model...")
def load_translation_model():
    # CPU, float32: NLLB-600M is ~2.4GB in memory. Loaded lazily, only the
    # first time a translation is actually requested.
    return hf_pipeline("translation", model=NLLB_MODEL, device=-1, torch_dtype=torch.float32)


def translate_text(text, target_lang):
    if not text or not text.strip():
        return ""
    try:
        translator = load_translation_model()
        result = translator(text, src_lang=PASHTO_CODE, tgt_lang=target_lang, max_length=512, num_beams=4)
        return result[0]["translation_text"]
    except Exception as e:
        return "Translation error: " + str(e)


def verify_translation(original_pashto, translated_text, target_lang):
    if not translated_text or "error" in translated_text.lower():
        return None, "Could not verify", ""
    try:
        translator = load_translation_model()
        back = translator(translated_text, src_lang=target_lang, tgt_lang=PASHTO_CODE, max_length=512, num_beams=4)
        back_text = back[0]["translation_text"]
        ratio = SequenceMatcher(None, original_pashto.strip(), back_text.strip()).ratio()
        score = round(ratio * 100, 1)
        if score >= 75:
            label = "✅ High confidence"
        elif score >= 50:
            label = "⚠️ Medium confidence"
        else:
            label = "❌ Low confidence — review recommended"
        return score, label, back_text
    except Exception as e:
        return None, f"Verification error: {e}", ""


def compute_wer(reference, hypothesis):
    if not JIWER_AVAILABLE or not reference.strip():
        return None
    try:
        return round(jiwer_wer(reference.strip(), hypothesis.strip()) * 100, 2)
    except Exception:
        return None

# ── Export helpers ────────────────────────────────────────────────────────

def make_csv_bytes(rows, headers):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for row in rows:
        w.writerow([str(cell) if cell is not None else "" for cell in row])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def make_excel_bytes(rows, headers):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transcriptions"
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        rtl_cols = {"Pashto", "Urdu", "Original", "Edited", "Reference"}
        for row_idx, row in enumerate(rows, 2):
            for col_idx, (header, cell_val) in enumerate(zip(headers, row), 1):
                val = str(cell_val) if cell_val is not None else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if any(rtl in header for rtl in rtl_cols):
                    cell.alignment = Alignment(horizontal="right")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    except ImportError:
        return make_csv_bytes(rows, headers)

# ── AI Chat — optional, external free APIs ─────────────────────────────────

def _try_hf_inference(prompt, system_msg, token):
    if not token:
        return None
    req_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    models_to_try = [
        ("HuggingFaceH4/zephyr-7b-beta", f"<|system|>\n{system_msg}</s>\n<|user|>\n{prompt}</s>\n<|assistant|>"),
        ("mistralai/Mistral-7B-Instruct-v0.1", f"[INST] {system_msg}\n\n{prompt} [/INST]"),
        ("google/flan-t5-large", f"{system_msg}\n\n{prompt}"),
    ]
    for model_id, formatted_prompt in models_to_try:
        url = f"https://api-inference.huggingface.co/models/{model_id}"
        payload = {
            "inputs": formatted_prompt,
            "parameters": {"max_new_tokens": 400, "temperature": 0.7, "do_sample": True, "return_full_text": False},
            "options": {"wait_for_model": True, "use_cache": False},
        }
        try:
            r = requests.post(url, headers=req_headers, json=payload, timeout=500)
            if r.status_code == 200:
                data = r.json()
                text = ""
                if isinstance(data, list) and data:
                    text = data[0].get("generated_text", "").strip()
                elif isinstance(data, dict):
                    text = data.get("generated_text", "").strip()
                if text and len(text) > 10:
                    return text
        except Exception:
            continue
    return None


def _try_pollinations(prompt, system_msg):
    try:
        import urllib.parse
        combined = f"{system_msg}\n\nUser question: {prompt}"
        encoded = urllib.parse.quote(combined[:800])
        r = requests.get(f"https://text.pollinations.ai/{encoded}", timeout=500)
        if r.status_code == 200 and r.text.strip():
            return r.text.strip()
    except Exception:
        pass
    return None


def ask_ai_free(prompt, system_msg, token):
    result = _try_hf_inference(prompt, system_msg, token)
    if result:
        return result, None
    result = _try_pollinations(prompt, system_msg)
    if result:
        return result, None
    return None, (
        "AI Assistant is unavailable right now.\n\n"
        "1. Double-check your Hugging Face token is valid (Settings → Change token below)\n"
        "2. Wait 30 seconds and try again (free models may be waking up)\n"
        "3. Check your internet connection"
    )

# ── UI helpers ───────────────────────────────────────────────────────────

def wer_badge(wer_score):
    if wer_score is None:
        return
    if wer_score <= 15:
        bg, label = "#22c55e", "Excellent"
    elif wer_score <= 35:
        bg, label = "#f59e0b", "Good"
    elif wer_score <= 60:
        bg, label = "#f97316", "Fair"
    else:
        bg, label = "#ef4444", "Poor"
    accuracy = max(0, round(100 - wer_score, 1))
    st.markdown(
        f'<div style="display:flex;gap:10px;align-items:center;margin:8px 0;">'
        f'<div style="background:{bg};color:#fff;font-weight:700;padding:5px 14px;border-radius:20px;font-size:0.85rem;">'
        f'WER: {wer_score}%</div>'
        f'<div style="background:#1e293b;color:#94a3b8;font-size:0.82rem;padding:5px 12px;border-radius:20px;">'
        f'~{accuracy}% accuracy — {label}</div></div>',
        unsafe_allow_html=True,
    )


def show_translation_box(tid, pashto_text, existing_en, existing_ur):
    st.markdown("**Translate:**")
    col_en, col_ur, col_both = st.columns(3)
    with col_en:
        if st.button("To English", key="en_" + str(tid), use_container_width=True):
            with st.spinner("Translating to English..."):
                en = translate_text(pashto_text, ENGLISH_CODE)
            save_translations(tid, en, existing_ur)
            st.session_state["t_en_" + str(tid)] = en
            st.rerun()
    with col_ur:
        if st.button("To Urdu", key="ur_" + str(tid), use_container_width=True):
            with st.spinner("Translating to Urdu..."):
                ur = translate_text(pashto_text, URDU_CODE)
            save_translations(tid, existing_en, ur)
            st.session_state["t_ur_" + str(tid)] = ur
            st.rerun()
    with col_both:
        if st.button("Both Languages", key="both_" + str(tid), use_container_width=True):
            with st.spinner("Translating..."):
                en = translate_text(pashto_text, ENGLISH_CODE)
                ur = translate_text(pashto_text, URDU_CODE)
            save_translations(tid, en, ur)
            st.session_state["t_en_" + str(tid)] = en
            st.session_state["t_ur_" + str(tid)] = ur
            st.rerun()

    en_val = st.session_state.get("t_en_" + str(tid), existing_en or "")
    ur_val = st.session_state.get("t_ur_" + str(tid), existing_ur or "")

    if en_val:
        st.text_area("English", value=en_val, height=85, key="disp_en_" + str(tid))
        col_dl_en, col_vfy_en = st.columns(2)
        with col_dl_en:
            st.download_button("Download English", data=en_val.encode("utf-8"), file_name="en_" + str(tid) + ".txt",
                                mime="text/plain; charset=utf-8", key="dl_en_" + str(tid))
        with col_vfy_en:
            if st.button("Verify English", key="btn_vfy_en_" + str(tid), use_container_width=True):
                with st.spinner("Back-translating to verify..."):
                    score, lbl, back = verify_translation(pashto_text, en_val, ENGLISH_CODE)
                st.session_state["vfy_en_result_" + str(tid)] = (score, lbl, back)
        vfy = st.session_state.get("vfy_en_result_" + str(tid))
        if vfy:
            score, lbl, back = vfy
            st.info(f"**Translation Quality:** {lbl} (similarity {score}%)")
            if back:
                with st.expander("Back-translated Pashto"):
                    st.markdown(f'<div class="pashto-text">{back}</div>', unsafe_allow_html=True)

    if ur_val:
        st.text_area("Urdu", value=ur_val, height=85, key="disp_ur_" + str(tid))
        col_dl_ur, col_vfy_ur = st.columns(2)
        with col_dl_ur:
            st.download_button("Download Urdu", data=ur_val.encode("utf-8"), file_name="ur_" + str(tid) + ".txt",
                                mime="text/plain; charset=utf-8", key="dl_ur_" + str(tid))
        with col_vfy_ur:
            if st.button("Verify Urdu", key="btn_vfy_ur_" + str(tid), use_container_width=True):
                with st.spinner("Back-translating to verify..."):
                    score, lbl, back = verify_translation(pashto_text, ur_val, URDU_CODE)
                st.session_state["vfy_ur_result_" + str(tid)] = (score, lbl, back)
        vfy = st.session_state.get("vfy_ur_result_" + str(tid))
        if vfy:
            score, lbl, back = vfy
            st.info(f"**Translation Quality:** {lbl} (similarity {score}%)")
            if back:
                with st.expander("Back-translated Pashto"):
                    st.markdown(f'<div class="pashto-text">{back}</div>', unsafe_allow_html=True)

# ── Pages ────────────────────────────────────────────────────────────────

def about_page():
    st.markdown("""
    <style>
    .about-hero{background:linear-gradient(135deg,#1e1b4b,#312e81);border-radius:16px;
        padding:2.5rem 2rem;margin-bottom:1.5rem;color:white;}
    .about-hero h1{font-size:2.2rem;font-weight:800;margin-bottom:0.3rem;}
    .about-hero p{font-size:1.05rem;color:#c7d2fe;}
    .feature-card{background:#1e293b;border:1px solid #334155;border-radius:12px;
        padding:1.2rem 1.4rem;margin-bottom:0.8rem;color:#e2e8f0;}
    .feature-card h4{color:#a78bfa;margin-bottom:0.3rem;font-size:1rem;}
    .feature-card p{font-size:0.88rem;color:#94a3b8;margin:0;}
    </style>
    <div class="about-hero">
        <h1>🎙️ Pashto Whisper STT</h1>
        <p>A fine-tuned Speech-to-Text system for Pakistani Pashto — runs locally on your
        own device, all data stored on this machine.</p>
    </div>
    """, unsafe_allow_html=True)

    st.subheader("What is this app?")
    st.markdown(
        "This application uses a **fine-tuned OpenAI Whisper** model trained on "
        "**Pakistani Pashto** speech, with LoRA weights fully merged into the base model. "
        "It converts spoken Pashto into written Pashto text, then optionally translates to "
        "English or Urdu. Transcription runs through the project's own Transcription API "
        "(started separately); translation and everything else runs locally in this app."
    )
    st.markdown("---")
    st.subheader("Features")
    features = [
        ("🎤 Multi-format Audio Upload", "WAV, MP3, M4A, OGG, FLAC, MP4, OPUS, WEBM, AAC, WMA."),
        ("📝 Pashto Transcription", "Merged Whisper model served by the project's own Transcription API."),
        ("✏️ Editable Transcriptions", "Correct mistakes and save the improved version."),
        ("📊 WER Scoring", "Paste reference text to get Word Error Rate + accuracy %."),
        ("🌐 English & Urdu Translation", "NLLB-200 600M runs locally — no API needed."),
        ("✅ Translation Verification", "Back-translates output and shows similarity score."),
        ("📁 CSV & Excel Export", "UTF-8 BOM encoding — Pashto/Urdu text displays correctly."),
        ("🔒 Secure Accounts", "Username + email + salted password hashing, private per-user data."),
    ]
    col1, col2 = st.columns(2)
    for i, (title, desc) in enumerate(features):
        target = col1 if i % 2 == 0 else col2
        with target:
            st.markdown(f'<div class="feature-card"><h4>{title}</h4><p>{desc}</p></div>', unsafe_allow_html=True)


def auth_page():
    st.markdown(
        '<style>.auth-title{font-size:2.4rem;font-weight:800;'
        'background:linear-gradient(135deg,#6366f1,#a78bfa);'
        '-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:0.2rem;}'
        '.auth-sub{color:#94a3b8;font-size:1rem;margin-bottom:2rem;}</style>'
        '<div class="auth-title">Pashto Whisper</div>'
        '<div class="auth-sub">Local Speech Transcription — your data stays on this device</div>',
        unsafe_allow_html=True,
    )
    tab_login, tab_signup = st.tabs(["Login", "Create Account"])
    with tab_login:
        with st.form("login_form"):
            identifier = st.text_input("Username or Email")
            password = st.text_input("Password", type="password")
            if st.form_submit_button("Login", use_container_width=True, type="primary"):
                ok, uid, uname = verify_user(identifier.strip(), password)
                if ok:
                    st.session_state.logged_in = True
                    st.session_state.user_id = uid
                    st.session_state.username = uname
                    st.rerun()
                else:
                    st.error("Invalid credentials.")
    with tab_signup:
        with st.form("signup_form"):
            new_user = st.text_input("Username")
            new_email = st.text_input("Email")
            new_pass = st.text_input("Password", type="password", help="At least 8 characters, letters + numbers")
            confirm = st.text_input("Confirm Password", type="password")
            if st.form_submit_button("Create Account", use_container_width=True, type="primary"):
                error = validate_signup(new_user.strip(), new_email.strip(), new_pass, confirm)
                if error:
                    st.error(error)
                else:
                    ok, msg = create_user(new_user.strip(), new_email.strip().lower(), new_pass)
                    st.success(msg) if ok else st.error(msg)


def transcribe_page():
    st.header("Transcribe Audio")
    st.caption("Supported: " + ", ".join("." + f for f in SUPPORTED_FORMATS))

    healthy, detail = check_api_health()
    if not healthy:
        st.error(
            f"Transcription API is unreachable at `{API_BASE_URL}` ({detail}). "
            "Start it first, e.g.: `uvicorn api.main:app --host 0.0.0.0 --port 8000`"
        )

    uploaded_files = st.file_uploader("Upload audio files", type=SUPPORTED_FORMATS, accept_multiple_files=True)
    ref_text = st.text_area("Reference text (optional) — paste correct Pashto text to calculate WER", height=70)
    if not uploaded_files:
        st.info("Upload audio files above to begin transcription.")
        return
    st.markdown(f"**{len(uploaded_files)} file(s) selected**")
    if not st.button("Transcribe All", type="primary", use_container_width=True):
        return

    progress_bar = st.progress(0, text="Starting...")
    all_results = []
    tmp_dir = Path("/tmp") if os.name != "nt" else APP_DATA_DIR / "_tmp"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    for i, f in enumerate(uploaded_files):
        st.markdown("---")
        st.markdown(f"**[{i+1}/{len(uploaded_files)}] {f.name}**")
        tmp_path = tmp_dir / f"upload_{i}_{f.name}"
        with open(tmp_path, "wb") as out:
            out.write(f.getbuffer())

        user_dir = AUDIO_STORE / str(st.session_state.user_id)
        user_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        saved_path = user_dir / f"{ts}_{f.name}"
        with open(saved_path, "wb") as out:
            out.write(f.getbuffer())

        st.audio(str(tmp_path))
        with st.spinner(f"Transcribing {f.name} via API..."):
            text, api_wer, proc_time, model_version, err = transcribe_via_api(str(tmp_path), ref_text)

        if err:
            st.error(err)
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            progress_bar.progress((i + 1) / len(uploaded_files), text=f"Failed {i+1}/{len(uploaded_files)}")
            continue

        wer_score = api_wer if api_wer is not None else compute_wer(ref_text, text)
        tid = save_transcription(
            st.session_state.user_id, f.name, text,
            audio_filepath=str(saved_path), reference=ref_text.strip() or None, wer_score=wer_score,
        )
        all_results.append((tid, f.name, text, wer_score))
        st.text_area(f"Pashto Transcription — {f.name}", value=text, height=100, key=f"res_{i}_{tid}")
        if model_version or proc_time is not None:
            st.caption(f"Model: {model_version or 'n/a'}  ·  API processing time: {proc_time if proc_time is not None else 'n/a'}s")
        if wer_score is not None:
            wer_badge(wer_score)
        elif ref_text.strip() and not JIWER_AVAILABLE:
            st.warning("Install jiwer to enable WER: pip install jiwer")
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        progress_bar.progress((i + 1) / len(uploaded_files), text=f"Completed {i+1}/{len(uploaded_files)}")

    st.markdown("---")
    st.success(f"All {len(uploaded_files)} file(s) transcribed and saved!")
    headers = ["#", "Filename", "Transcription", "WER (%)", "Saved At"]
    rows = [[idx, fname, text, ws or "", datetime.now().isoformat()] for idx, (tid, fname, text, ws) in enumerate(all_results, 1)]
    col_csv, col_xlsx = st.columns(2)
    with col_csv:
        st.download_button("Download Session as CSV", data=make_csv_bytes(rows, headers),
                            file_name="session_transcriptions.csv", mime="text/csv; charset=utf-8")
    with col_xlsx:
        st.download_button("Download Session as Excel", data=make_excel_bytes(rows, headers),
                            file_name="session_transcriptions.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def history_page():
    st.header("My Transcription History")
    rows = get_user_history(st.session_state.user_id)
    if not rows:
        st.info("No transcriptions yet. Head to Transcribe to get started.")
        return
    st.markdown(f"**{len(rows)} transcription(s) on record**")

    export_headers = ["ID", "Filename", "Original", "Edited", "Reference", "WER (%)", "English", "Urdu", "Audio Path", "Date"]
    export_rows = [list(r) for r in rows]

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.download_button("Export All as CSV", data=make_csv_bytes(export_rows, export_headers),
                            file_name="all_transcriptions.csv", mime="text/csv; charset=utf-8", use_container_width=True)
    with col_b:
        st.download_button("Export All as Excel", data=make_excel_bytes(export_rows, export_headers),
                            file_name="all_transcriptions.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with col_c:
        lines = []
        for r in rows:
            tid, fname, orig, edited, ref, ws, en_t, ur_t, audio_fp, ts = r
            lines += ["=" * 60, f"File   : {fname}", f"Date   : {ts[:16].replace('T', ' ')}",
                      f"WER    : {str(ws) + '%' if ws is not None else 'N/A'}", f"Pashto :\n{edited or orig or ''}"]
            if en_t:
                lines.append(f"English:\n{en_t}")
            if ur_t:
                lines.append(f"Urdu   :\n{ur_t}")
            lines.append("")
        st.download_button("Export All as TXT", data="\n".join(lines).encode("utf-8"),
                            file_name="all_transcriptions.txt", mime="text/plain; charset=utf-8", use_container_width=True)
    st.markdown("---")

    for r in rows:
        tid, filename, original, edited, reference, wer_score, en_t, ur_t, audio_filepath, created_at = r
        date_str = created_at[:16].replace("T", " ") if created_at else ""
        label = f"{filename}  |  {date_str}"
        if wer_score is not None:
            label += f"  |  WER {wer_score}%"
        with st.expander(label):
            st.caption(f"Record #{tid}")
            if audio_filepath and os.path.exists(audio_filepath):
                st.audio(audio_filepath, format="audio/*", start_time=0)
            else:
                st.warning(f"Audio file not found: {filename}")
            if original and original != (edited or original):
                with st.expander("View original (unedited)"):
                    st.markdown(f'<div class="pashto-text">{original}</div>', unsafe_allow_html=True)
            edited_val = st.text_area("Pashto Transcription (editable)", value=edited or original or "",
                                       key=f"edit_{tid}", height=120)
            new_ref = st.text_input("Reference Pashto text for WER", value=reference or "", key=f"ref_{tid}")
            if wer_score is not None:
                wer_badge(wer_score)
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                if st.button("Save Edits", key=f"save_{tid}", use_container_width=True, type="primary"):
                    new_wer = compute_wer(new_ref, edited_val) if new_ref.strip() else None
                    update_transcription(tid, edited_val, new_ref.strip() or None, new_wer)
                    st.success("Saved!")
                    st.rerun()
            with col2:
                dk = f"confirm_delete_{tid}"
                if st.session_state.get(dk, False):
                    st.warning("Are you sure?")
                    if st.button("Confirm Delete", key=f"confirm_del_{tid}", use_container_width=True, type="primary"):
                        delete_transcription(tid, audio_filepath)
                        st.session_state[dk] = False
                        st.rerun()
                    if st.button("Cancel", key=f"cancel_del_{tid}", use_container_width=True):
                        st.session_state[dk] = False
                        st.rerun()
                else:
                    if st.button("Remove", key=f"remove_{tid}", use_container_width=True, type="primary"):
                        st.session_state[dk] = True
                        st.rerun()
            with col3:
                st.download_button("Download TXT", data=edited_val.encode("utf-8"), file_name=f"{filename}.txt",
                                    mime="text/plain; charset=utf-8", key=f"dl_t_{tid}", use_container_width=True)
            with col4:
                sh = ["Filename", "Pashto", "English", "Urdu", "WER (%)", "Date"]
                sr = [[filename, edited_val, en_t or "", ur_t or "", wer_score or "", created_at]]
                st.download_button("Download CSV", data=make_csv_bytes(sr, sh), file_name=f"{filename}.csv",
                                    mime="text/csv; charset=utf-8", key=f"dl_c_{tid}", use_container_width=True)
            st.markdown("---")
            show_translation_box(tid, edited_val, en_t, ur_t)


def _token_setup_form(user_id, first_time=True):
    """Prompt to save a Hugging Face token for this account."""
    if first_time:
        st.info(
            "**One-time setup:** the AI Assistant uses your own free Hugging Face token "
            "so it can call HF's inference API on your behalf. It's saved to your account "
            "on this device — you won't need to enter it again unless it expires or you "
            "want to change it."
        )
    st.markdown(
        "Don't have one? Create a free token at "
        "[huggingface.co/settings/tokens](https://huggingface.co/settings/tokens) "
        "(a **read** token is enough), then paste it below."
    )
    with st.form("hf_token_form", clear_on_submit=False):
        token_input = st.text_input("Hugging Face token", type="password", placeholder="hf_...")
        submitted = st.form_submit_button("Save Token", type="primary", use_container_width=True)
    if submitted:
        token_input = token_input.strip()
        if not token_input:
            st.warning("Please paste a token first.")
        elif not token_input.startswith("hf_"):
            st.warning("That doesn't look like a Hugging Face token — it should start with `hf_`.")
        else:
            save_user_ai_token(user_id, token_input)
            st.success("Token saved! You won't need to enter it again on this device.")
            st.rerun()


def ai_chat_page():
    st.header("AI Assistant")
    st.caption("Ask anything about your Pashto transcriptions — powered by your own free Hugging Face token.")

    user_id = st.session_state.user_id
    token = get_user_ai_token(user_id)

    if not token:
        _token_setup_form(user_id, first_time=True)
        return

    masked = f"hf_...{token[-4:]}" if len(token) > 4 else "hf_***"
    with st.expander(f"Using saved token ({masked}) — change or remove it"):
        _token_setup_form(user_id, first_time=False)
        if st.button("Remove saved token", use_container_width=True):
            clear_user_ai_token(user_id)
            st.success("Token removed.")
            st.rerun()

    rows = get_user_history(user_id)
    trans_labels = ["(none — type your own context)"]
    trans_texts = [""]
    for r in rows:
        tid, fname, orig, edited, ref, ws, en_t, ur_t, audio_fp, ts = r
        trans_labels.append(f"{fname} | {ts[:10]}")
        trans_texts.append(edited or orig or "")
    idx = st.selectbox("Load a saved transcription as context", range(len(trans_labels)), format_func=lambda i: trans_labels[i])
    context_text = st.text_area("Context (Pashto text)", value=trans_texts[idx], height=130)
    system_msg = st.text_input("System instruction",
                                value="You are a helpful assistant specializing in Pashto language and transcription analysis.")
    user_question = st.text_area("Your question", height=100)
    if st.button("Ask AI", type="primary", use_container_width=True):
        if not user_question.strip():
            st.warning("Please type a question first.")
        else:
            prompt = (f"Transcription context:\n{context_text.strip()}\n\n" if context_text.strip() else "") + user_question.strip()
            with st.spinner("Asking AI (up to 60 seconds on free tier)..."):
                answer, err = ask_ai_free(prompt, system_msg, token)
            if err:
                st.error(err)
            else:
                st.markdown("---")
                st.markdown("**AI Response:**")
                st.markdown(answer)
                st.download_button("Download AI Response",
                                    data=f"Question:\n{user_question}\n\nContext:\n{context_text}\n\nAnswer:\n{answer}".encode("utf-8"),
                                    file_name="ai_response.txt", mime="text/plain; charset=utf-8")

# ── Main ────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="Pashto Whisper", page_icon="🎙", layout="wide", initial_sidebar_state="expanded")
    st.markdown(
        '<style>'
        '[data-testid="stSidebar"]{background:#0f172a;}'
        '[data-testid="stSidebar"] *{color:#e2e8f0 !important;}'
        # Urdu font — Jameel Noori Nastaleeq (local static/ file first, CDN fallback)
        "@font-face {"
        "  font-family: 'JameelNooriNastaleeq';"
        "  src: url('app/static/JameelNooriNastaleeq.woff2') format('woff2'),"
        "       url('https://cdn.jsdelivr.net/gh/tariq-abdullah/urdu-web-font-CDN/JameelNooriNastaleeq.woff') format('woff');"
        "  font-weight: normal; font-style: normal; font-display: swap;"
        "}"
        # Pashto font — Noto Naskh Arabic from Google Fonts. Unlike Urdu Nastaleeq
        # fonts, this has correct glyph forms for Pashto-only letters
        # (ږ ښ ړ ډ ټ ڼ etc.) instead of rendering them broken/disconnected.
        "@import url('https://fonts.googleapis.com/css2?family=Noto+Naskh+Arabic:wght@400;700&display=swap');"
        # Pashto fields
        'textarea[aria-label*="Pashto"], input[aria-label*="Pashto"] {'
        "  font-family: 'Noto Naskh Arabic', serif !important;"
        "  font-size: 22px !important; line-height: 2 !important;"
        "  direction: rtl; text-align: right;"
        "}"
        ".pashto-text {"
        "  font-family: 'Noto Naskh Arabic', serif;"
        "  font-size: 22px; line-height: 2; direction: rtl; text-align: right;"
        "}"
        # Urdu fields (translation output only)
        'textarea[aria-label="Urdu"] {'
        "  font-family: 'JameelNooriNastaleeq', 'Noto Nastaliq Urdu', serif !important;"
        "  font-size: 22px !important; line-height: 2.1 !important;"
        "  direction: rtl; text-align: right;"
        "}"
        ".urdu-text {"
        "  font-family: 'JameelNooriNastaleeq', 'Noto Nastaliq Urdu', serif;"
        "  font-size: 22px; line-height: 2.1; direction: rtl; text-align: right;"
        "}"
        "</style>",
        unsafe_allow_html=True,
    )
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if not st.session_state.logged_in:
        auth_page()
        return

    with st.sidebar:
        st.markdown("## 🎙️ Pashto Whisper")
        st.markdown(f"**User: {st.session_state.username}**")
        st.markdown("---")
        nav_options = ["About", "Transcribe", "History", "AI Assistant"]
        page = st.radio("Navigate", nav_options, label_visibility="collapsed")
        st.markdown("---")
        if not JIWER_AVAILABLE:
            st.warning("WER disabled. Run: pip install jiwer")
        device_label = "GPU (CUDA)" if torch.cuda.is_available() else "CPU"
        st.caption(f"Local device : {device_label} (translation)")
        api_healthy, api_detail = check_api_health()
        if api_healthy:
            info = get_api_model_info()
            model_label = info.get("model", "STT-Whisper-Pashto") if info else "connected"
            st.caption(f"Transcription API : ✅ {model_label}")
        else:
            st.caption(f"Transcription API : ❌ unreachable ({api_detail})")
        st.caption(f"Storage     : {APP_DATA_DIR}")
        ai_status = "configured" if get_user_ai_token(st.session_state.user_id) else "needs setup"
        st.caption(f"AI Assistant: {ai_status}")
        st.markdown("---")
        if st.button("Free Up Memory", use_container_width=True,
                      help="Unloads cached models from RAM. They'll reload next time you use them."):
            st.cache_resource.clear()
            st.success("Cached models cleared from memory.")
        if st.button("Logout", use_container_width=True):
            for k in ("logged_in", "user_id", "username"):
                st.session_state[k] = None
            st.session_state.logged_in = False
            st.rerun()

    if page == "About":
        about_page()
    elif page == "Transcribe":
        transcribe_page()
    elif page == "History":
        history_page()
    elif page == "AI Assistant":
        ai_chat_page()


if __name__ == "__main__":
    main()